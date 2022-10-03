import pessoa as p
from openpyxl import load_workbook

#define a classe FUNCIONARIO importando a classe Pessoa da lib pessoa (pessoa.Pessoa)

class Funcionario(p.Pessoa):
    def __init__(self, cpf, nome, setor, salario, entrada, saida):
        super().__init__(cpf, nome)
        self.setor = setor
        self.salario = salario
        self.entrada = entrada
        self.saida = saida
        

#definição dos CARGOS DA EMPRESA herdando a classe Funcionario:


#ADMINISTRADOR

class Administrador(Funcionario):
    def __init__(self, cpf=None, nome=None, setor=None, salario=5000, entrada=None, saida=None, ajudaDeCusto=None, remuneracao=None):
        super().__init__(cpf, nome, setor, salario, entrada, saida)
        self.ajudaDeCusto = ajudaDeCusto
        self.remuneracao = remuneracao
        
    def contratar(self, i):

        j = str(len(i))
        #abre planilha para edição
        planilha = load_workbook("./cadastro.xlsx")
        tabela = planilha.active

        print("Insira os dados do novo Administrador:")

        self.nome = str(input("Nome: "))
        self.cpf = str(input("CPF: "))
        self.setor = str(input("Setor: "))
        self.entrada = str(input("Entrada: "))
        self.saida = str(input("Saída: "))
        self.salario = input("Salario combinado: ")
        self.ajudaDeCusto = input("Valor do vale gasolina: ")
        
        #Insere dados na planilha iniciando da linha 2:
        tabela['A'+ j].value = self.nome # A1 - nome
        tabela['B'+ j].value = self.cpf # B1 - cpf
        tabela['C'+ j].value = self.setor # C1 - setor
        tabela['D'+ j].value = self.entrada # D1 - entrada
        tabela['E'+ j].value = self.saida # E1 - saida
        tabela['F'+ j].value = "Administrador" # F1 - Cargo
        tabela['G'+ j].value = self.salario # G1 - salario
        tabela['H'+ j].value = self.ajudaDeCusto # H1 - ajudaDeCusto
        tabela['I'+ j].value = 0 # I1 - valorVendas
        tabela['J'+ j].value = 0 #J1 - valorProducao
        tabela['K'+ j].value = 0 # K1 - comissao

        self.remuneracao = float(self.salario) + float(self.ajudaDeCusto)
        tabela['L'+ j].value = self.remuneracao # L1 - remuneracao

        planilha.save("./cadastro.xlsx")
        print(f"\n\n {tabela['F'+ j].value} cadastrado com sucesso!\n\n")




# VENDEDOR

class Vendedor(Funcionario):
    def __init__(self, cpf=None, nome=None, setor="Loja", salario=None, entrada=None, saida=None, valorVendas=0, comissao=30, remuneracao=None):
        super().__init__(cpf, nome, setor, salario, entrada, saida)
        self.valorVendas = valorVendas
        self.comissao = comissao
        self.remuneracao = remuneracao

    def contratar(self, i):

        j = str(len(i))

        #abre planilha para edição
        planilha = load_workbook("./cadastro.xlsx")
        tabela = planilha.active

        print("Insira os dados do novo Vendedor:")
        
        self.nome = str(input("Nome: "))
        self.cpf = str(input("CPF: "))
        self.setor = str(input("Setor: "))
        self.entrada = str(input("Entrada: "))
        self.saida = str(input("Saída: "))
        self.salario = input("Salario combinado: ")
        
        #Insere dados na planilha iniciando da linha 2:
        tabela['A'+ j].value = self.nome # A1 - nome
        tabela['B'+ j].value = self.cpf # B1 - cpf
        tabela['C'+ j].value = self.setor # C1 - setor
        tabela['D'+ j].value = self.entrada # D1 - entrada
        tabela['E'+ j].value = self.saida # E1 - saida
        tabela['F'+ j].value = "Vendedor" # F1 - Cargo
        tabela['G'+ j].value = self.salario # G1 - salario
        tabela['H'+ j].value = 0 # H1 - ajudaDeCusto
        tabela['I'+ j].value = 0 # I1 - valorVendas
        tabela['J'+ j].value = 0 # J1 - valorProducao
        tabela['K'+ j].value = 0.3 # K1 - comissao de 30%

        self.remuneracao = float(self.salario) + float(self.valorVendas * self.comissao)
        tabela['L'+ j].value = self.remuneracao # L1 - remuneracao

        planilha.save("./cadastro.xlsx")
        print(f"\n\n {tabela['F'+ j].value} cadastrado com sucesso!\n\n")




#OPERARIO

class Operario(Funcionario):
    def __init__(self, cpf=None, nome=None, setor="Fabrica", salario=2000, entrada=None, saida=None, valorProducao=0, comissao=20, remuneracao=None):
        super().__init__(cpf, nome, setor, salario, entrada, saida)
        self.valorProducao = valorProducao
        self.comissao = comissao
        self.remuneracao = remuneracao


    def contratar(self,i):

        j = str(len(i))

        #abre planilha para edição
        planilha = load_workbook("./cadastro.xlsx")
        tabela = planilha.active

        print("Insira os dados do novo Operário:")

        self.nome = str(input("Nome: "))
        self.cpf = str(input("CPF: "))
        self.setor = str(input("Setor: "))
        self.entrada = str(input("Entrada: "))
        self.saida = str(input("Saída: "))
        self.salario = input("Salario combinado: ")

        #Insere dados na planilha iniciando da linha 2:
        tabela['A'+ j].value = self.nome # A1 - nome
        tabela['B'+ j].value = self.cpf # B1 - cpf
        tabela['C'+ j].value = self.setor # C1 - setor
        tabela['D'+ j].value = self.entrada # D1 - entrada
        tabela['E'+ j].value = self.saida # E1 - saida
        tabela['F'+ j].value = "Operario" # F1 - Cargo
        tabela['G'+ j].value = self.salario # G1 - salario
        tabela['H'+ j].value = 0# H1 - ajudaDeCusto
        tabela['I'+ j].value = 0 # I1 - valorVendas
        tabela['J'+ j].value = 0 # J1 - valorProducao
        tabela['K'+ j].value = 0.2 # K1 - comissao de 20%

        self.renumeracao = float(self.salario) + float(self.valorProducao * self.comissao)
        tabela['L'+ j].value = self.remuneracao # L1 - remuneracao

        planilha.save("./cadastro.xlsx")
        print(f"\n\n {tabela['F'+ j].value} cadastrado com sucesso!\n\n")


