from openpyxl import load_workbook
from funcionario import Administrador, Vendedor, Operario
from menu import Menu

Menu.titulo(None)

while True:
    op = Menu.principal(None)
    if op == "1":
        
        i = [1] # este i serve para controlar as tabelas do excel, a cada laço, desce uma tabela
        while True:
            cargo = Menu.cargo(None)
            if cargo == "0":
                break

            else:
                i.append(1) #adiciona "1" na lista para contabilizar registros na planilha
                if cargo == "1":
                    vendedor = Vendedor()
                    vendedor.contratar(i)
                    
                elif cargo == "2":
                    operario = Operario()
                    operario.contratar(i)
                    
                elif cargo == "3":
                    adm = Administrador()
                    adm.contratar(i)

    elif op =="2":
        planilha = load_workbook("./cadastro.xlsx")
        tabela = planilha.active
        for i in range(2,tabela.max_row+1):
            print("-"*50)
            print("NOME: {}".format(tabela['A'+ str(i)].value),end="   ")
            print("CPF: {}".format(tabela['B'+ str(i)].value),end="   ")
            print("SETOR: {}".format(tabela['C'+ str(i)].value))
            print("ENTRADA: {}".format(tabela['D'+ str(i)].value),end="   ")
            print("SAÍDA: {}".format(tabela['E'+ str(i)].value), end="    ")
            print("CARGO: {}".format(tabela['F'+ str(i)].value))
            print("SALÁRIO: {}".format(tabela['G'+ str(i)].value),end="   ")
            print("AUXÍLIO: {}".format(tabela['H'+ str(i)].value),end="   ")
            print("VENDAS: {}".format(tabela['I'+ str(i)].value))
            print("PRODUÇÃO: {}".format(tabela['J'+ str(i)].value),end="   ")
            print("COMISSÃO: {}".format(tabela['K'+ str(i)].value),end="   ")
            print("REMUNERAÇÃO: {}".format(tabela['L'+ str(i)].value))
            
    else:
        print("\n**** Opção inválida")
                




