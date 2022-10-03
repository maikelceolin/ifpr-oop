import pessoa as p
from openpyxl import load_workbook

#define a classe empresa
class Empresa():
    def __init__(self, nome, cnpj):
        self.nome = nome
        self.cnpj = cnpj


#define a classe FUNCIONARIO importando a classe Pessoa da lib pessoa (pessoa.Pessoa)

class Funcionario(p.Pessoa):
    def __init__(self, cpf, nome, setor, salario, entrada, saida):
        super().__init__(cpf, nome)
        self.setor = setor
        self.salario = salario
        self.entrada = entrada
        self.saida = saida
        

#definição dos CARGOS DA EMPRESA herdando a classe Funcionario:


class Administrador(Funcionario):
    def __init__(self, cpf=None, nome=None, setor=None, salario=5000, entrada=None, saida=None, auxilio=None, remuneracao=None):
        super().__init__(cpf, nome, setor, salario, entrada, saida)
        self.auxilio = auxilio
        self.remuneracao = remuneracao
        
    def contratar(self, i):

        j = str(len(i))
        #abre planilha para edição
        planilha = load_workbook("./cadastro.xlsx")
        tabela = planilha.active

        print("Insira os dados do novo Administrador:")

        self.nome = str(input("Nome: "))
        self.cpf = str(input("CPF: "))
        self.setor = str(input("Setor: "))
        self.entrada = str(input("Entrada: "))
        self.saida = str(input("Saída: "))
        self.salario = input("Salario combinado: ")
        self.auxilio = input("Valor do vale gasolina: ")
        
        #Insere dados na planilha iniciando da linha 2:
        tabela['A'+ j].value = self.nome # A1 - nome
        tabela['B'+ j].value = self.cpf # B1 - cpf
        tabela['C'+ j].value = self.setor # C1 - setor
        tabela['D'+ j].value = self.entrada # D1 - entrada
        tabela['E'+ j].value = self.saida # E1 - saida
        tabela['F'+ j].value = "Administrador" # F1 - Cargo
        tabela['G'+ j].value = self.salario # G1 - salario
        tabela['H'+ j].value = self.auxilio # H1 - auxilio
        tabela['I'+ j].value = 0 # I1 - vendas
        tabela['J'+ j].value = 0 #J1 - producao
        tabela['K'+ j].value = 0 # K1 - comissao

        self.remuneracao = float(self.salario) + float(self.auxilio)
        tabela['L'+ j].value = self.remuneracao # L1 - remuneracao

        planilha.save("./cadastro.xlsx")

class Vendedor(Funcionario):
    def __init__(self, cpf=None, nome=None, setor="Loja", salario=None, entrada=None, saida=None, vendas=0, comissao=30, remuneracao=None):
        super().__init__(cpf, nome, setor, salario, entrada, saida)
        self.vendas = vendas
        self.comissao = comissao
        self.remuneracao = remuneracao

    def contratar(self, i):

        j = str(len(i))

        #abre planilha para edição
        planilha = load_workbook("./cadastro.xlsx")
        tabela = planilha.active

        print("Insira os dados do novo Vendedor:")
        
        self.nome = str(input("Nome: "))
        self.cpf = str(input("CPF: "))
        self.setor = str(input("Setor: "))
        self.entrada = str(input("Entrada: "))
        self.saida = str(input("Saída: "))
        self.salario = input("Salario combinado: ")
        
        #Insere dados na planilha iniciando da linha 2:
        tabela['A'+ j].value = self.nome # A1 - nome
        tabela['B'+ j].value = self.cpf # B1 - cpf
        tabela['C'+ j].value = self.setor # C1 - setor
        tabela['D'+ j].value = self.entrada # D1 - entrada
        tabela['E'+ j].value = self.saida # E1 - saida
        tabela['F'+ j].value = "Vendedor" # F1 - Cargo
        tabela['G'+ j].value = self.salario # G1 - salario
        tabela['H'+ j].value = 0 # H1 - auxilio
        tabela['I'+ j].value = 0 # I1 - vendas
        tabela['J'+ j].value = 0 # J1 - producao
        tabela['K'+ j].value = 0.3 # K1 - comissao de 30%

        self.remuneracao = float(self.salario) + float(self.vendas * self.comissao)
        tabela['L'+ j].value = self.remuneracao # L1 - remuneracao

        planilha.save("./cadastro.xlsx")

class Operario(Funcionario):
    def __init__(self, cpf=None, nome=None, setor="Fabrica", salario=2000, entrada=None, saida=None, producao=0, comissao=20, remuneracao=None):
        super().__init__(cpf, nome, setor, salario, entrada, saida)
        self.producao = producao
        self.comissao = comissao
        self.remuneracao = remuneracao


    def contratar(self,i):

        j = str(len(i))

        #abre planilha para edição
        planilha = load_workbook("./cadastro.xlsx")
        tabela = planilha.active

        print("Insira os dados do novo Operário:")

        self.nome = str(input("Nome: "))
        self.cpf = str(input("CPF: "))
        self.setor = str(input("Setor: "))
        self.entrada = str(input("Entrada: "))
        self.saida = str(input("Saída: "))
        self.salario = input("Salario combinado: ")

        #Insere dados na planilha iniciando da linha 2:
        tabela['A'+ j].value = self.nome # A1 - nome
        tabela['B'+ j].value = self.cpf # B1 - cpf
        tabela['C'+ j].value = self.setor # C1 - setor
        tabela['D'+ j].value = self.entrada # D1 - entrada
        tabela['E'+ j].value = self.saida # E1 - saida
        tabela['F'+ j].value = "Operario" # F1 - Cargo
        tabela['G'+ j].value = self.salario # G1 - salario
        tabela['H'+ j].value = 0# H1 - auxilio
        tabela['I'+ j].value = 0 # I1 - vendas
        tabela['J'+ j].value = 0 # J1 - producao
        tabela['K'+ j].value = 0.2 # K1 - comissao de 20%

        self.renumeracao = float(self.salario) + float(self.producao * self.comissao)
        tabela['L'+ j].value = self.remuneracao # L1 - remuneracao

        planilha.save("./cadastro.xlsx")


#definição dos FORNECEDORES:

class Fornecedor(p.Pessoa):
    def __init__(self, cpf, nome, credito, divida):
        p.Pessoa.__init__(cpf, nome)
        self.credito = credito
        self.divida = divida

    def obterSaldo(self):
        saldo = self.credito - self.divida
        return saldo